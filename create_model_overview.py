#!/usr/bin/env python3
"""
Erstellt eine Excel-Übersicht über alle trainierten Modelle und deren Datensätze.
"""

import os
import yaml
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
except ImportError:
    print("pandas nicht installiert. Installiere mit: pip install pandas openpyxl")
    exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl nicht installiert. Installiere mit: pip install openpyxl")
    exit(1)


def find_hydra_yamls(outputs_dir: str) -> list:
    """Findet alle hydra.yaml Dateien im outputs-Verzeichnis (nur Hauptordner, nicht .hydra/)."""
    hydra_files = []
    for root, dirs, files in os.walk(outputs_dir):
        # Überspringe .hydra/ Unterordner
        if ".hydra" in root:
            continue
        if "hydra.yaml" in files:
            hydra_files.append(os.path.join(root, "hydra.yaml"))
    return sorted(hydra_files)


def extract_dataset_name(data_path: str) -> str:
    """Extrahiert den Datensatz-Namen aus dem Pfad."""
    if data_path:
        return os.path.basename(data_path.rstrip('/'))
    return "N/A"


def parse_dataset_info(dataset_name: str) -> dict:
    """Parst Informationen aus dem Datensatz-Namen (Namenskonvention)."""
    info = {
        "n_episodes": "N/A",
        "robot_opacity": "N/A",
        "n_primitives": "N/A",
        "n_cameras": "N/A",
        "n_cubes": "N/A",
        "fusion_type": "N/A",
        "flags": ""
    }
    
    parts = dataset_name.split('_')
    flags = []
    
    for part in parts:
        if part.startswith("NEps"):
            info["n_episodes"] = part[4:]
        elif part.startswith("RobOpac"):
            info["robot_opacity"] = part[7:]
        elif part.startswith("NPrim"):
            info["n_primitives"] = part[5:]
        elif part.startswith("NCams"):
            info["n_cameras"] = part[5:]
        elif part.startswith("NCube"):
            info["n_cubes"] = part[5:]
        elif part.startswith("Fuse"):
            info["fusion_type"] = part[4:]
        elif part in ["EEFfix", "TrackGrip", "ChannelStack", "GrGrDiff", "SterileBg"]:
            flags.append(part)
        elif part.startswith("ActInt"):
            info["n_primitives"] = f"ActInt{part[6:]}"
    
    info["flags"] = ", ".join(flags)
    return info


def extract_model_info(hydra_path: str) -> dict:
    """Extrahiert alle relevanten Informationen aus einer hydra.yaml."""
    try:
        with open(hydra_path, 'r') as f:
            cfg = yaml.safe_load(f)
    except Exception as e:
        print(f"Fehler beim Lesen von {hydra_path}: {e}")
        return None
    
    # Modellpfad extrahieren
    saved_folder = cfg.get("saved_folder", "")
    model_name = "/".join(saved_folder.split("/")[-2:]) if saved_folder else "N/A"
    
    # Datensatz-Pfad extrahieren
    data_path = ""
    if "env" in cfg and "dataset" in cfg["env"]:
        data_path = cfg["env"]["dataset"].get("data_path", "")
    
    dataset_name = extract_dataset_name(data_path)
    dataset_info = parse_dataset_info(dataset_name)
    
    # Training-Parameter
    training = cfg.get("training", {})
    
    return {
        "Modell": model_name,
        "Datensatz": dataset_name,
        "Data Path": data_path,
        "Epochs": training.get("epochs", "N/A"),
        "Batch Size": training.get("batch_size", "N/A"),
        "num_hist": cfg.get("num_hist", "N/A"),
        "num_pred": cfg.get("num_pred", "N/A"),
        "frameskip": cfg.get("frameskip", "N/A"),
        "action_dim": cfg.get("env", {}).get("action_dim", "N/A"),
        "proprio_dim": cfg.get("env", {}).get("proprio_dim", "N/A"),
        "action_emb_dim": cfg.get("action_emb_dim", "N/A"),
        "proprio_emb_dim": cfg.get("proprio_emb_dim", "N/A"),
        "Encoder LR": training.get("encoder_lr", "N/A"),
        "Decoder LR": training.get("decoder_lr", "N/A"),
        "Predictor LR": training.get("predictor_lr", "N/A"),
        "img_size": cfg.get("img_size", "N/A"),
        "DS: N Episodes": dataset_info["n_episodes"],
        "DS: Robot Opacity": dataset_info["robot_opacity"],
        "DS: N Primitives": dataset_info["n_primitives"],
        "DS: N Cameras": dataset_info["n_cameras"],
        "DS: N Cubes": dataset_info["n_cubes"],
        "DS: Fusion Type": dataset_info["fusion_type"],
        "DS: Flags": dataset_info["flags"],
        "W&B Run ID": cfg.get("wandb_run_id", "N/A"),
    }


def main():
    outputs_dir = Path(__file__).parent / "outputs"
    
    print(f"Suche hydra.yaml Dateien in: {outputs_dir}")
    hydra_files = find_hydra_yamls(str(outputs_dir))
    print(f"Gefunden: {len(hydra_files)} Modelle\n")
    
    # Alle Modell-Informationen sammeln
    models = []
    for hf in hydra_files:
        info = extract_model_info(hf)
        if info:
            models.append(info)
            print(f"  ✓ {info['Modell']}")
    
    if not models:
        print("Keine Modelle gefunden!")
        return
    
    # DataFrame erstellen
    df = pd.DataFrame(models)
    
    # Nach Modellname sortieren (chronologisch)
    df = df.sort_values("Modell", ascending=True)
    
    # Excel-Datei speichern
    output_file = outputs_dir.parent / "MODELL_UEBERSICHT.xlsx"
    
    # Mit Formatierung speichern
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Modelle', index=False)
        
        # Worksheet formatieren
        worksheet = writer.sheets['Modelle']
        
        # Spaltenbreiten anpassen
        for idx, col in enumerate(df.columns):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            # Maximale Breite begrenzen
            max_len = min(max_len, 50)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = max_len
        
        # Header fett
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
    
    print(f"\n✓ Excel-Datei gespeichert: {output_file}")
    print(f"  Anzahl Modelle: {len(models)}")


if __name__ == "__main__":
    main()
